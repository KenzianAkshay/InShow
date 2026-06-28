import csv
import io
import json
import os
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse

from app.auth import require_user
from app.db import DATABASE_PATH, connect
from app.ontology import (
    build_graph,
    class_instances,
    export_jsonld,
    infer_combined_ontology,
    read_ontology,
    schema_ontology,
)

DATA_DIR = Path(
    os.getenv("DATA_DIR") or Path(DATABASE_PATH).resolve().parent / "uploads"
)

router = APIRouter(prefix="/api", dependencies=[Depends(require_user)])


def _filled(row) -> int:
    """Count of non-blank cells in a row."""
    return sum(1 for c in row if c is not None and str(c).strip() != "")


def _detect_header_idx(rows: list) -> int:
    """Index of the most likely header row, skipping leading banner/title rows.

    Spreadsheets exported for people often start with a merged title/banner
    (e.g. "EXHIBITOR BOOTH AVAILABILITY & PRICING — TechExpo 2026") and maybe a
    blank line, with the real column headers a few rows down. A banner is narrow
    (one or two populated cells) while the header spans the table width, so we
    take the first row that is wide relative to the table — the header, since
    data follows it. No banner → row 0 wins, matching the old behaviour.
    """
    scan = rows[:25]
    width = max((_filled(r) for r in scan), default=0)
    if width <= 1:
        return 0
    threshold = max(2, (width * 3 + 4) // 5)  # ~0.6 of the table width
    for i, row in enumerate(scan):
        if _filled(row) >= threshold:
            return i
    return 0


def _records_from_rows(rows: list) -> list[dict]:
    """Turn a sheet's raw rows into records: find the header row (past any
    banner), de-duplicate/auto-name header cells, then map each data row.
    All cells coerced to strings, keeping Neo4j property types simple."""
    rows = [tuple(r) for r in rows]
    if not rows:
        return []
    start = _detect_header_idx(rows)
    header = rows[start]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(header):
        name = "" if h is None else str(h).strip()
        if not name:
            name = f"col{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    records = []
    for row in rows[start + 1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        records.append(
            {
                headers[i]: "" if val is None else str(val)
                for i, val in enumerate(row)
                if i < len(headers)
            }
        )
    return records


def _sheet_to_records(ws) -> list[dict]:
    return _records_from_rows(list(ws.iter_rows(values_only=True)))


def _parse_xlsx_sheets(content: bytes) -> dict[str, list[dict]]:
    """Every non-empty worksheet/tab becomes its own record set, keyed by tab
    name. The ontology builder combines them into one graph."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        records = _sheet_to_records(ws)
        if records:
            sheets[ws.title] = records
    return sheets


def parse_sheets(content: bytes, filename: str) -> dict[str, list[dict]]:
    """Parse a data source into one or more named record sets. Excel workbooks
    yield one set per tab; CSV/JSON yield a single set named after the file."""
    name = (filename or "").lower()
    stem = Path(filename or "data").stem or "data"
    if name.endswith((".xlsx", ".xlsm")):
        sheets = _parse_xlsx_sheets(content)
    elif name.endswith(".json"):
        data = json.loads(content.decode("utf-8-sig"))
        if isinstance(data, dict):
            data = [data]
        sheets = {stem: [d for d in data if isinstance(d, dict)]}
    elif name.endswith(".csv"):
        rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        sheets = {stem: _records_from_rows(rows)}
    else:
        raise HTTPException(400, "Unsupported file type; use .csv, .json, or .xlsx")
    sheets = {k: v for k, v in sheets.items() if v}
    if not sheets:
        raise HTTPException(400, "No records found in file")
    return sheets


def parse_records(content: bytes, filename: str) -> list[dict]:
    """Flat view of every record across all tabs (used for record counts)."""
    return [r for records in parse_sheets(content, filename).values() for r in records]


def _neo4j(request: Request):
    driver = request.app.state.neo4j
    try:
        driver.verify_connectivity()
    except Exception:
        raise HTTPException(503, "Neo4j is not available")
    return driver


@router.get("/data-sources")
def list_data_sources(project_id: int | None = None):
    conn = connect()
    if project_id is None:
        rows = conn.execute(
            "SELECT * FROM data_sources ORDER BY created_at DESC"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM data_sources WHERE show_project_id = ? "
            "ORDER BY created_at DESC",
            (project_id,),
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/data-sources/{source_id}")
def get_data_source(source_id: int):
    conn = connect()
    row = conn.execute(
        "SELECT * FROM data_sources WHERE id = ?", (source_id,)
    ).fetchone()
    conn.close()
    if row is None:
        raise HTTPException(404, "Data source not found")
    return dict(row)


@router.post("/data-sources", status_code=201)
async def create_data_source(
    project_id: int, file: UploadFile = File(...)
):
    content = await file.read()
    records = parse_records(content, file.filename or "data")
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()

    conn = connect()
    if conn.execute(
        "SELECT 1 FROM show_projects WHERE id = ?", (project_id,)
    ).fetchone() is None:
        conn.close()
        raise HTTPException(404, "Show project not found")
    cur = conn.execute(
        "INSERT INTO data_sources (name, type, status, show_project_id) "
        "VALUES (?, ?, ?, ?)",
        (file.filename or "data", ext, "ingested", project_id),
    )
    source_id = cur.lastrowid
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = DATA_DIR / f"{source_id}_{file.filename or 'data'}"
    path.write_bytes(content)
    conn.execute(
        "UPDATE data_sources SET location = ? WHERE id = ?",
        (str(path), source_id),
    )
    conn.commit()
    row = conn.execute(
        "SELECT * FROM data_sources WHERE id = ?", (source_id,)
    ).fetchone()
    conn.close()
    return {**dict(row), "records": len(records)}


@router.post("/data-sources/{source_id}/build")
def build_ontology(source_id: int, request: Request):
    conn = connect()
    row = conn.execute(
        "SELECT * FROM data_sources WHERE id = ?", (source_id,)
    ).fetchone()
    conn.close()
    if row is None:
        raise HTTPException(404, "Data source not found")
    if not row["location"] or not Path(row["location"]).exists():
        raise HTTPException(400, "Data source file is missing")

    content = Path(row["location"]).read_bytes()
    sheets = parse_sheets(content, row["name"])
    spec = infer_combined_ontology(sheets, Path(row["name"]).stem)

    driver = _neo4j(request)
    summary = build_graph(driver, spec, source_id, row["show_project_id"])

    conn = connect()
    conn.execute(
        "UPDATE data_sources SET status = 'in_graph' WHERE id = ?", (source_id,)
    )
    conn.commit()
    conn.close()
    return summary


@router.get("/ontology")
def get_ontology(request: Request, project_id: int):
    return read_ontology(_neo4j(request), project_id)


@router.get("/ontology/schema")
def get_ontology_schema(request: Request, project_id: int):
    """Class-level schema graph — the legible default view for the ontology layer."""
    return schema_ontology(_neo4j(request), project_id)


@router.get("/ontology/instances")
def get_ontology_instances(
    request: Request, project_id: int, label: str, limit: int = 150
):
    """Instances of one class plus their neighbours — drill-down from the schema."""
    return class_instances(_neo4j(request), project_id, label, limit)


@router.get("/ontology/export")
def export_ontology(request: Request, project_id: int, scope: str = "full"):
    """Download the project's ontology as JSON-LD. scope=schema for the model only
    (classes + relationship types) or full to include all instances."""
    scope = "schema" if scope == "schema" else "full"
    data = export_jsonld(_neo4j(request), project_id, scope)
    filename = f"ontology-project-{project_id}-{scope}.jsonld"
    return JSONResponse(
        content=data,
        media_type="application/ld+json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
