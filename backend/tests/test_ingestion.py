CSV = "exhibitor,booth,city\nAcme,A12,Berlin\nGlobex,B07,Munich\n"


def test_upload_csv(auth, project):
    r = auth.post(
        f"/api/data-sources?project_id={project}",
        files={"file": ("x.csv", CSV, "text/csv")},
    )
    assert r.status_code == 201
    body = r.json()
    assert body["records"] == 2
    assert body["type"] == "csv"
    assert body["status"] == "ingested"
    assert body["show_project_id"] == project


def test_upload_json(auth, project):
    data = '[{"a": 1}, {"a": 2}, {"a": 3}]'
    r = auth.post(
        f"/api/data-sources?project_id={project}",
        files={"file": ("x.json", data, "application/json")},
    )
    assert r.status_code == 201
    assert r.json()["records"] == 3


def test_upload_xlsx(auth, project):
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["exhibitor", "city", "booths"])
    ws.append(["Acme", "Berlin", 3])
    ws.append(["Globex", "Munich", 5])
    buf = io.BytesIO()
    wb.save(buf)

    r = auth.post(
        f"/api/data-sources?project_id={project}",
        files={
            "file": (
                "book.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 201
    assert r.json()["records"] == 2
    assert r.json()["type"] == "xlsx"


def test_upload_xlsx_multiple_tabs(auth, project):
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Exhibitors"
    ws1.append(["exhibitor", "city"])
    ws1.append(["Acme", "Berlin"])
    ws1.append(["Globex", "Munich"])
    ws2 = wb.create_sheet("Booths")
    ws2.append(["booth", "hall"])
    ws2.append(["A12", "H1"])
    buf = io.BytesIO()
    wb.save(buf)

    r = auth.post(
        f"/api/data-sources?project_id={project}",
        files={
            "file": (
                "multi.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 201
    # records counts rows across every tab (2 exhibitors + 1 booth)
    assert r.json()["records"] == 3


def test_build_combined_ontology_across_tabs(auth_up, project):
    """A real multi-tab workbook flows through parse -> combined inference ->
    graph build, with the City dimension shared across both tabs collapsing to a
    single node linked from each tab."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Exhibitors"
    ws1.append(["exhibitor", "city"])
    ws1.append(["Acme", "Berlin"])
    ws1.append(["Globex", "Berlin"])
    ws2 = wb.create_sheet("Sessions")
    ws2.append(["session", "city"])
    ws2.append(["Keynote", "Berlin"])
    ws2.append(["Workshop", "Berlin"])
    buf = io.BytesIO()
    wb.save(buf)

    sid = auth_up.post(
        f"/api/data-sources?project_id={project}",
        files={
            "file": (
                "expo.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    ).json()["id"]

    summary = auth_up.post(f"/api/data-sources/{sid}/build").json()
    # Exhibitors + Sessions + City; HAS_CITY shared; 4 record nodes + 1 City;
    # 4 edges (two records per tab -> the one shared City:Berlin).
    assert summary["classes"] == 3
    assert summary["relations"] == 1
    assert summary["nodes"] == 5
    assert summary["edges"] == 4


def test_records_from_rows_skips_banner_row():
    from app.ingestion import _records_from_rows

    rows = [
        ("EXHIBITOR BOOTH AVAILABILITY & PRICING — TechExpo 2026", None, None, None),
        (None, None, None, None),
        ("Booth ID", "Exhibitor Name", "Booth Type", "Balance Due"),
        ("BTH-1", "Acme", "Inline", "1200"),
        ("BTH-2", "Globex", "Corner", "0"),
    ]
    records = _records_from_rows(rows)
    assert len(records) == 2
    assert set(records[0]) == {"Booth ID", "Exhibitor Name", "Booth Type", "Balance Due"}
    assert records[0]["Exhibitor Name"] == "Acme"
    # the banner title must not leak in as a column
    assert all("TechExpo" not in k for r in records for k in r)


def test_records_from_rows_no_banner_uses_first_row():
    from app.ingestion import _records_from_rows

    rows = [("a", "b"), ("1", "2"), ("3", "4")]
    records = _records_from_rows(rows)
    assert records == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]


def test_records_from_rows_names_blank_and_duplicate_headers():
    from app.ingestion import _records_from_rows

    rows = [("Name", None, "Name"), ("Acme", "x", "y")]
    records = _records_from_rows(rows)
    assert set(records[0]) == {"Name", "col1", "Name_1"}


def test_upload_xlsx_with_title_banner(auth, project):
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["EXHIBITOR BOOTH PRICING — TechExpo 2026"])  # banner
    ws.append([None, None, None])  # blank
    ws.append(["Booth ID", "Exhibitor Name", "Balance Due"])  # real header
    ws.append(["BTH-1", "Acme", "1200"])
    ws.append(["BTH-2", "Globex", "0"])
    buf = io.BytesIO()
    wb.save(buf)

    r = auth.post(
        f"/api/data-sources?project_id={project}",
        files={
            "file": (
                "pricing.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 201
    assert r.json()["records"] == 2


def test_upload_unknown_project(auth):
    assert auth.post(
        "/api/data-sources?project_id=999999",
        files={"file": ("x.csv", CSV, "text/csv")},
    ).status_code == 404


def test_upload_bad_type(auth, project):
    assert auth.post(
        f"/api/data-sources?project_id={project}", files={"file": ("x.txt", "hi")}
    ).status_code == 400


def test_upload_empty(auth, project):
    assert auth.post(
        f"/api/data-sources?project_id={project}",
        files={"file": ("x.csv", "col1,col2\n")},
    ).status_code == 400


def test_list_and_get(auth, project):
    sid = auth.post(
        f"/api/data-sources?project_id={project}",
        files={"file": ("y.csv", CSV, "text/csv")},
    ).json()["id"]
    assert any(d["id"] == sid for d in auth.get("/api/data-sources").json())
    scoped = auth.get(f"/api/data-sources?project_id={project}").json()
    assert any(d["id"] == sid for d in scoped)
    assert auth.get(f"/api/data-sources/{sid}").json()["type"] == "csv"
    assert auth.get("/api/data-sources/999999").status_code == 404


def test_ontology_schema_and_instances_endpoints(auth_up, project):
    # MockDriver returns no rows -> empty but well-formed payloads.
    schema = auth_up.get(f"/api/ontology/schema?project_id={project}").json()
    assert schema == {"classes": [], "edges": []}
    inst = auth_up.get(
        f"/api/ontology/instances?project_id={project}&label=City"
    ).json()
    assert inst == {"nodes": [], "edges": []}


def test_build_and_ontology_require_neo4j(auth, project):
    sid = auth.post(
        f"/api/data-sources?project_id={project}",
        files={"file": ("z.csv", CSV, "text/csv")},
    ).json()["id"]
    # neo4j is unreachable in the hermetic env
    assert auth.post(f"/api/data-sources/{sid}/build").status_code == 503
    assert auth.get(f"/api/ontology?project_id={project}").status_code == 503
    assert auth.post("/api/data-sources/999999/build").status_code == 404
